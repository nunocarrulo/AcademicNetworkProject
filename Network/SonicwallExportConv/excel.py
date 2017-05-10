
import openpyxl
import collections
import re


class excelHelper:

    def __init__(self, rules, addrObjects, addrGroups, serviceObjects, serviceGroups):
        self.rules = rules
        self.oAddrObjects = collections.OrderedDict(sorted(addrObjects.items()))
        self.addrGroups = addrGroups
        self.oServiceObjects = collections.OrderedDict(sorted(serviceObjects.items()))
        self.serviceGroups = serviceGroups

    def createWorkbook(self):
        print ('Building File...')
        rowID=2
        print ('Creating Workbook...')
        self.wb = openpyxl.Workbook()
        print ('Creating Policy Sheet...')
        #start workbook with a new sheet
        sheetPolicy = self.wb.create_sheet(index=0, title='Policy')
        #Write Policy Sheet Header
        sheetPolicy.cell(row=1, column=1, value="Source Zone")
        sheetPolicy.cell(row=1, column=2, value="Dest Zone")
        sheetPolicy.cell(row=1, column=3, value="Source Net")
        sheetPolicy.cell(row=1, column=4, value="Dest Net")
        sheetPolicy.cell(row=1, column=5, value="Dest Service")
        sheetPolicy.cell(row=1, column=6, value="Action")
        sheetPolicy.cell(row=1, column=7, value="Status")
        sheetPolicy.cell(row=1, column=8, value="Comment")

        #aux variables
        prevSrcZone=""
        prevDestZone=""

        for x in self.rules:
            if x["ruleSrcZone"] != prevSrcZone or x["ruleDestZone"] != prevDestZone:
                #if is not the first row, adds an empty one
                if rowID != 2:
                    rowID += 1
                #write ZONE to ZONE header
                zoneToZoneHeader = x["ruleSrcZone"] + " to " + x["ruleDestZone"]
                sheetPolicy.cell(row=rowID, column=1, value=zoneToZoneHeader)
                rowID += 1
            #else:
            sheetPolicy.cell(row=rowID, column=1, value=x["ruleSrcZone"])
            sheetPolicy.cell(row=rowID, column=2, value=x["ruleDestZone"])
            sheetPolicy.cell(row=rowID, column=3, value=x["ruleSrcNet"])
            sheetPolicy.cell(row=rowID, column=4, value=x["ruleDestNet"])
            sheetPolicy.cell(row=rowID, column=5, value=x["ruleDestService"])
            sheetPolicy.cell(row=rowID, column=6, value=x["ruleAction"])
            sheetPolicy.cell(row=rowID, column=7, value=x["ruleStatus"])
            if re.match('^Auto added', x["ruleComment"]) or re.match('^Auto-added', x["ruleComment"]):
                x["ruleComment"] = "Auto added rule"
            sheetPolicy.cell(row=rowID, column=8, value=x["ruleComment"])
            rowID += 1
            prevSrcZone=x["ruleSrcZone"]
            prevDestZone=x["ruleDestZone"]

        print ('Creating Address Object Sheet...')
        #create second sheet
        sheetAddrObj = self.wb.create_sheet(index=1, title='Address Objects')
        rowID=2
        #Write Address Objects Sheet Header
        sheetAddrObj.cell(row=1, column=1, value="Address Name")
        sheetAddrObj.cell(row=1, column=2, value="Zone")
        sheetAddrObj.cell(row=1, column=3, value="IP")
        sheetAddrObj.cell(row=1, column=4, value="Subnet")

        for addr, addrFields in self.oAddrObjects.items():
            #Check if address Object is not a Group (Group addrType==8)
            if addrFields["addrType"] == '1':
                sheetAddrObj.cell(row=rowID, column=1, value=addr)
                sheetAddrObj.cell(row=rowID, column=2, value=addrFields["addrZone"])
                sheetAddrObj.cell(row=rowID, column=3, value=addrFields["addrIP"])
                sheetAddrObj.cell(row=rowID, column=4, value=addrFields["addrSubnet"])
                rowID += 1

        print ('Creating Address Group Sheet...')
        sheetAddrGrp = self.wb.create_sheet(index=2, title='Address Groups')
        rowID=2
        #Write Address Groups Header
        sheetAddrGrp.cell(row=1, column=1, value="Group Name")
        sheetAddrGrp.cell(row=1, column=2, value="Object Name")

        for group, groupObjects in self.addrGroups.items():
            #write group name
            sheetAddrGrp.cell(row=rowID, column=1, value=group)
            for groupObj in groupObjects:
                #write group objects
                sheetAddrGrp.cell(row=rowID, column=2, value=groupObj)
                rowID += 1
            rowID += 1


        print ('Creating Service Objects Sheet...')
        #create fourth sheet
        sheetSrvObj = self.wb.create_sheet(index=3, title='Service Objects')
        rowID=2
        #Write Service Objects Sheet Header
        sheetSrvObj.cell(row=1, column=1, value="Service Name")
        sheetSrvObj.cell(row=1, column=2, value="Start Port")
        sheetSrvObj.cell(row=1, column=3, value="End Port")
        sheetSrvObj.cell(row=1, column=4, value="Protocol")
        sheetSrvObj.cell(row=1, column=5, value="Object Type")
        #write services objects
        for service, serviceFields in self.oServiceObjects.items():
            sheetSrvObj.cell(row=rowID, column=1, value=service)
            sheetSrvObj.cell(row=rowID, column=2, value=serviceFields["serviceStartPort"])
            sheetSrvObj.cell(row=rowID, column=3, value=serviceFields["serviceEndPort"])
            sheetSrvObj.cell(row=rowID, column=4, value=serviceFields["serviceProtocol"])
            sheetSrvObj.cell(row=rowID, column=5, value=serviceFields["serviceType"])
            rowID += 1

        print ('Creating Service Group Sheet...')
        #create fifth sheet
        sheetSrvGrp = self.wb.create_sheet(index=4, title='Service Groups')
        rowID=2
        #Write Service Group sheet Header
        sheetSrvGrp.cell(row=1, column=1, value="Group Name")
        sheetSrvGrp.cell(row=1, column=2, value="Service Name")
        #write service Groups
        for serviceGroup,serviceGroupObjects in self.serviceGroups.items():
            sheetSrvGrp.cell(row=rowID, column=1, value=serviceGroup)
            for serviceObj in serviceGroupObjects:
                sheetSrvGrp.cell(row=rowID, column=2, value=serviceObj)
                rowID += 1
            rowID += 1


    def writeData (self, filename):
        print ('Writing file: ' + filename)
        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))

        #save workbook file
        self.wb.save(filename)
        print('Done')